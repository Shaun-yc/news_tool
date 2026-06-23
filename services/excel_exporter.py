from datetime import datetime
import io
import re
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _sanitize(value):
    if isinstance(value, str):
        return _ILLEGAL_XML_CHARS.sub("", value)
    return value


HEADERS = [
    "doc_id",
    "headline",
    "responsible_unit",
    "compiled_by",
    "project",
    "functional_cat",
    "subcategory",
    "pubdate",
    "source",
    "source_url",
    "content",
    "content_tran",
    "attach_cnt",
    "attach_fname",
]


def extract_source(url):
    try:
        hostname = urlparse(url).hostname or ""
        domain = hostname.removeprefix("www.")
        return domain.split(".")[0].upper() or "LINK"
    except Exception:
        return "LINK"


def format_date(raw_date):
    try:
        return datetime.strptime(raw_date[:10], "%Y-%m-%d").strftime("%Y/%m/%d")
    except (TypeError, ValueError):
        return raw_date


def build_excel_report(news_list, week_date):
    """Build the standard 14-column Excel report and return an in-memory file."""
    excel_buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "週新聞"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for column, header in enumerate(HEADERS, 1):
        cell = worksheet.cell(row=1, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for index, news in enumerate(news_list):
        en_title = news.get("en_title", "")
        headline = f"{news['zh_title']}\n{en_title}" if en_title else news["zh_title"]
        row_data = [
            f"{week_date}_{index + 1:02d}",
            headline,
            "淨零推動組",
            "永智顧問",
            "碳市場國際合作與企業能力建構計畫",
            "排放管理;國際事務;調適韌性;減量交易",
            news.get("subcategory", ""),
            format_date(news.get("pubdate", "")),
            extract_source(news.get("source_url", "")),
            news.get("source_url", ""),
            news.get("content", ""),
            news.get("en_content", ""),
            0,
            "",
        ]
        for column, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=index + 2, column=column, value=_sanitize(value))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [18, 40, 14, 12, 30, 30, 35, 12, 15, 45, 60, 60, 10, 12]
    for column, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    worksheet.freeze_panes = "A2"
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

