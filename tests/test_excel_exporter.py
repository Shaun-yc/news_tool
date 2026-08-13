import unittest

import openpyxl

from services.excel_exporter import build_excel_report, extract_source, format_date


class ExcelExporterTests(unittest.TestCase):
    def test_extract_source_uses_first_domain_segment(self):
        self.assertEqual(extract_source("https://www.example.com/news/1"), "EXAMPLE")
        self.assertEqual(extract_source(""), "LINK")

    def test_format_date_uses_report_format(self):
        self.assertEqual(format_date("2026-06-01"), "2026/06/01")
        self.assertEqual(format_date("unknown"), "unknown")

    def test_build_excel_report_keeps_standard_columns(self):
        report = build_excel_report(
            [
                {
                    "zh_title": "中文標題",
                    "en_title": "English title",
                    "subcategory": "排放管理;國際事務",
                    "pubdate": "2026-06-01",
                    "source_url": "https://www.example.com/news/1",
                    "content": "中文摘要",
                    "en_content": "English content",
                }
            ],
            "20260601",
        )
        workbook = openpyxl.load_workbook(report)
        worksheet = workbook["週新聞"]

        self.assertEqual(worksheet.max_column, 14)
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet["A2"].value, "20260601_01")
        self.assertEqual(worksheet["B2"].value, "中文標題\nEnglish title")
        self.assertEqual(worksheet["C2"].value, "碳費推動組")
        self.assertEqual(worksheet["H2"].value, "2026/06/01")
        self.assertEqual(worksheet["I2"].value, "EXAMPLE")

    def test_build_excel_report_limits_visible_height_for_full_article_text(self):
        article = "Complete article text. " * 1000
        report = build_excel_report(
            [
                {
                    "zh_title": "Long article",
                    "source_url": "https://www.example.com/news/1",
                    "content": "Summary",
                    "en_content": article,
                }
            ],
            "20260731",
        )
        workbook = openpyxl.load_workbook(report)
        worksheet = workbook["週新聞"]

        self.assertEqual(worksheet["L2"].value, article)
        self.assertEqual(worksheet.row_dimensions[2].height, 60)

    def test_build_excel_report_marks_empty_content_tran_for_manual_review(self):
        report = build_excel_report(
            [
                {
                    "zh_title": "Article with missing source content",
                    "source_url": "https://www.example.com/news/1",
                    "content": "Chinese summary",
                    "en_content": "",
                }
            ],
            "20260731",
        )
        workbook = openpyxl.load_workbook(report)
        worksheet = workbook["週新聞"]

        self.assertEqual(worksheet["L2"].value, "來源內文為空，請手動確認來源網址。")

    def test_build_excel_report_treats_whitespace_content_tran_as_empty(self):
        report = build_excel_report(
            [
                {
                    "zh_title": "Article with whitespace source content",
                    "source_url": "https://www.example.com/news/1",
                    "content": "Chinese summary",
                    "en_content": "  \n\t  ",
                }
            ],
            "20260731",
        )
        workbook = openpyxl.load_workbook(report)
        worksheet = workbook["週新聞"]

        self.assertEqual(worksheet["L2"].value, "來源內文為空，請手動確認來源網址。")

    def test_build_excel_report_matches_complete_two_news_result_baseline(self):
        report = build_excel_report(
            [
                {
                    "zh_title": "第一篇中文標題",
                    "en_title": "First English title",
                    "subcategory": "排放管理;減量交易",
                    "pubdate": "2026-06-01",
                    "source_url": "https://www.example.com/news/1",
                    "content": "第一篇中文摘要",
                    "en_content": "First English article",
                },
                {
                    "zh_title": "第二篇中文標題",
                    "subcategory": "調適韌性",
                    "pubdate": "not-a-date",
                    "source_url": "https://news.example.org/path/2",
                    "content": "第二篇中文摘要",
                    "en_content": "",
                },
            ],
            "20260601",
        )
        worksheet = openpyxl.load_workbook(report).active

        self.assertEqual(worksheet.title, "週新聞")
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            [
                "doc_id",
                "headline",
                "responsible_unit",
                "compiled_by",
                "project",
                "functional_cat",
                "subcategory",
                "pubdate",
                "source",
                "source_url",
                "content",
                "content_tran",
                "attach_cnt",
                "attach_fname",
            ],
        )
        self.assertEqual(
            [
                [cell.value for cell in row]
                for row in worksheet.iter_rows(min_row=2, max_row=3, max_col=14)
            ],
            [
                [
                    "20260601_01",
                    "第一篇中文標題\nFirst English title",
                    "碳費推動組",
                    "永智顧問",
                    "碳市場國際合作與企業能力建構計畫",
                    "排放管理;國際事務;調適韌性;減量交易",
                    "排放管理;減量交易",
                    "2026/06/01",
                    "EXAMPLE",
                    "https://www.example.com/news/1",
                    "第一篇中文摘要",
                    "First English article",
                    0,
                    None,
                ],
                [
                    "20260601_02",
                    "第二篇中文標題",
                    "碳費推動組",
                    "永智顧問",
                    "碳市場國際合作與企業能力建構計畫",
                    "排放管理;國際事務;調適韌性;減量交易",
                    "調適韌性",
                    "not-a-date",
                    "NEWS",
                    "https://news.example.org/path/2",
                    "第二篇中文摘要",
                    "來源內文為空，請手動確認來源網址。",
                    0,
                    None,
                ],
            ],
        )
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(
            [worksheet.column_dimensions[chr(65 + index)].width for index in range(14)],
            [18, 40, 14, 12, 30, 30, 35, 12, 15, 45, 60, 60, 10, 12],
        )
        self.assertEqual(
            [worksheet.row_dimensions[index].height for index in (2, 3)],
            [60, 60],
        )

        header = worksheet["A1"]
        self.assertEqual(header.fill.fill_type, "solid")
        self.assertEqual(header.fill.fgColor.type, "rgb")
        self.assertEqual(header.fill.fgColor.rgb, "001F4E79")
        self.assertTrue(header.font.bold)
        self.assertEqual(header.font.color.type, "rgb")
        self.assertEqual(header.font.color.rgb, "00FFFFFF")
        self.assertEqual(header.font.sz, 10)
        self.assertEqual(header.alignment.horizontal, "center")
        self.assertTrue(header.alignment.wrap_text)

        data_cell = worksheet["A2"]
        self.assertTrue(data_cell.alignment.wrap_text)
        self.assertEqual(data_cell.alignment.vertical, "top")


if __name__ == "__main__":
    unittest.main()

